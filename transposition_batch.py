import pandas as pd
import openpyxl
import sys
import glob

def process(software_output):
    xls=pd.read_excel(software_output)

    select_column=['组分名','峰面积[%]']
    xls_selected=xls[select_column]
    xls_selected=xls_selected.dropna(subset=['组分名'])

    lines=xls_selected.to_numpy()
    my_dict={}
    exist=['异戊烷','23DMB','2MP','3MP','24DMP','223TMB','2MH','23DMP','3MH','224TMP','25DMH','223TMP','24DMH','234TMP','233TMP','33DMH','23DMH','225TMH']
    for row in lines:
        if row[0].replace(',', '').replace('-', '') in exist:
            my_dict[row[0].replace(',', '').replace('-', '')]=row[1]

    for components in exist:
        try:
            my_dict[components]
        except KeyError:
            my_dict[components]=0

    sorted_dict = dict(sorted(my_dict.items(), key=lambda x: exist.index(x[0])))

       
    workbook = openpyxl.Workbook()
    sheet = workbook.active

# 将字典的键存储到第一行
    keys = list(sorted_dict.keys())
    for col_num, key in enumerate(keys, 1):
        sheet.cell(row=1, column=col_num+2, value=key)

# 将字典的值存储到第二行
    values = list(sorted_dict.values())
    for col_num, value in enumerate(values, 1):
        sheet.cell(row=2, column=col_num+2, value=value)
    sheet.cell(row=2,column=1,value=software_output[-8:-6]+':'+software_output[-6:-4])
# 保存Excel文件
    workbook.save(software_output+'_processed.xlsx')
#process('BF2-0124-1932.xls')

if __name__ == "__main__":
    # 检查是否有足够的参数（文件路径）
    excel_files = glob.glob("BF2-*.xls")
    for xlss in excel_files:
        process(xlss)

    sys.exit()
